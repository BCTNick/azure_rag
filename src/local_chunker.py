from __future__ import annotations

import base64
import html
import io
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from typing import Any

from flask import Flask, jsonify, request
from openpyxl import load_workbook

ARTICLE_RE = re.compile(r"(?m)^\s*(?:[Aa]rticle|[Aa]rt\.)\s+(\d+[a-zA-Z\-]*)\b(?!\s*\()(?:\s+[A-Z][^\n]*)?$")
CHAPTER_RE = re.compile(r"(?im)^\s*chapter\s+([ivxlcdm0-9]+)\s*$")
ANNEX_RE = re.compile(r"(?im)^\s*annex\s+([ivxlcdm0-9a-zA-Z\-]+)\b.*$")
_HTML_SCRIPT_STYLE_RE = re.compile(r"(?is)<(script|style)[^>]*>.*?</\1>")
_HTML_TAG_RE = re.compile(r"(?s)<[^>]+>")
_HTML_SPACE_RE = re.compile(r"[ \t]+")
_HTML_NEWLINE_RE = re.compile(r"\n{3,}")
_STRUCTURAL_BLOCK_RE = re.compile(r"^(art|rct)_(\d+[a-zA-Z\-]*)$")
_CHAPTER_ID_RE = re.compile(r"^cpt_([ivxlcdm0-9]+)(?:\.|$)", re.IGNORECASE)
_SPACED_ARTICLE_RE = re.compile(r"\bA\s*r\s*t\s*i\s*c\s*l\s*e\b", re.IGNORECASE)
_SPACED_CHAPTER_RE = re.compile(r"\bC\s*h\s*a\s*p\s*t\s*e\s*r\b", re.IGNORECASE)
_SPACED_ANNEX_RE = re.compile(r"\bA\s*n\s*n\s*e\s*x\b", re.IGNORECASE)
_INLINE_PDF_ARTICLE_RE = re.compile(
    r"(?<!\n)\s+(Article\s+\d+[a-zA-Z\-]*\s+(?!of\b|thereof\b|point\b)[A-Z][^\n]{1,80})"
)
_VOID_HTML_TAGS = {"area", "base", "br", "col", "embed", "hr", "img", "input", "link", "meta", "param", "source", "track", "wbr"}
_HTML_SEPARATOR_TAGS = {"br", "p", "div", "li", "tr", "td", "th", "table", "section"}

app = Flask(__name__)


@dataclass
class StructuredLegalBlock:
    kind: str
    number: str
    chapter_num: int
    text: str


def _safe_int(value: str, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _roman_to_int(value: str, default: int = 0) -> int:
    raw = value.strip().upper()
    if not raw:
        return default
    if raw.isdigit():
        return int(raw)

    values = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    total = 0
    previous = 0
    for char in reversed(raw):
        current = values.get(char)
        if current is None:
            return default
        if current < previous:
            total -= current
        else:
            total += current
            previous = current
    return total if total > 0 else default


def _clean_text(text: str) -> str:
    text = html.unescape(text)
    text = text.replace("\u00a0", " ")
    text = text.replace("\r", "\n")
    text = _HTML_SPACE_RE.sub(" ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = _HTML_NEWLINE_RE.sub("\n\n", text)
    return text.strip()


def _annotate_chunk(page_row_num: int, chapter_num: int, article_num: str, annex_num: str, body: str) -> str:
    return (
        f"[page_row_num={page_row_num} | chapter_num={chapter_num} | "
        f"article_num={article_num} | annex_num={annex_num}]\n{body.strip()}"
    )


def _prefix_split_chunks(text: str, chunk_size: int = 2200, overlap: int = 220) -> list[str]:
    chunks = _fallback_chunk(text, chunk_size=chunk_size, overlap=overlap)
    if len(chunks) <= 1:
        return chunks

    first_line = chunks[0].splitlines()[0].strip()
    if not re.match(r"(?i)^(article|recital)\s+", first_line):
        return chunks

    prefixed: list[str] = []
    for idx, chunk in enumerate(chunks):
        if idx == 0 or chunk.startswith(first_line):
            prefixed.append(chunk)
        else:
            prefixed.append(f"{first_line}\n{chunk}")
    return prefixed


class EurLexHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.blocks: list[StructuredLegalBlock] = []
        self.in_doc_html = False
        self.doc_depth = 0
        self.current_chapter_num = 0
        self.current_kind: str | None = None
        self.current_number = ""
        self.current_block_chapter_num = 0
        self.current_block_depth = 0
        self.current_parts: list[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attr_map = {key.lower(): value or "" for key, value in attrs}
        element_id = attr_map.get("id", "")

        if not self.in_doc_html:
            if tag == "div" and element_id == "docHtml":
                self.in_doc_html = True
                self.doc_depth = 1
            return

        if tag not in _VOID_HTML_TAGS:
            self.doc_depth += 1

        if tag in {"script", "style"}:
            self.skip_depth += 1
            return

        chapter_match = _CHAPTER_ID_RE.match(element_id)
        if chapter_match:
            self.current_chapter_num = _roman_to_int(chapter_match.group(1), default=self.current_chapter_num)

        block_match = _STRUCTURAL_BLOCK_RE.match(element_id)
        if self.current_kind is None and tag == "div" and block_match:
            self.current_kind = "article" if block_match.group(1) == "art" else "recital"
            self.current_number = block_match.group(2)
            self.current_block_chapter_num = self.current_chapter_num if self.current_kind == "article" else 0
            self.current_block_depth = 1
            self.current_parts = []
            return

        if self.current_kind is not None:
            if tag in _HTML_SEPARATOR_TAGS:
                self.current_parts.append("\n")
            if tag not in _VOID_HTML_TAGS:
                self.current_block_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if self.skip_depth:
            if tag in {"script", "style"}:
                self.skip_depth -= 1
            return

        if self.current_kind is not None:
            if tag in _HTML_SEPARATOR_TAGS:
                self.current_parts.append("\n")
            if tag not in _VOID_HTML_TAGS:
                self.current_block_depth -= 1
                if self.current_block_depth == 0:
                    self._finish_block()

        if self.in_doc_html and tag not in _VOID_HTML_TAGS:
            self.doc_depth -= 1
            if self.doc_depth <= 0:
                self.in_doc_html = False

    def handle_data(self, data: str) -> None:
        if self.current_kind is None or self.skip_depth:
            return
        if data.strip():
            self.current_parts.append(data)

    def _finish_block(self) -> None:
        text = _clean_text(" ".join(self.current_parts))
        if text:
            heading = f"{self.current_kind.capitalize()} {self.current_number}"
            if not text.lower().startswith(heading.lower()):
                text = f"{heading}\n{text}"
            self.blocks.append(
                StructuredLegalBlock(
                    kind=self.current_kind,
                    number=self.current_number,
                    chapter_num=self.current_block_chapter_num,
                    text=text,
                )
            )

        self.current_kind = None
        self.current_number = ""
        self.current_block_chapter_num = 0
        self.current_block_depth = 0
        self.current_parts = []


def _split_pages(text: str) -> list[str]:
    if "\f" in text:
        pages = [p.strip() for p in text.split("\f")]
        pages = [p for p in pages if p]
        if pages:
            return pages
    return [text.strip()] if text.strip() else []


def _fallback_chunk(text: str, chunk_size: int = 2200, overlap: int = 220) -> list[str]:
    if not text.strip():
        return []

    step = max(1, chunk_size - overlap)
    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = min(len(text), start + chunk_size)
        piece = text[start:end].strip()
        if piece:
            chunks.append(piece)
        if end >= len(text):
            break
        start += step
    return chunks


def _article_number_as_int(value: str) -> int | None:
    match = re.match(r"^(\d+)", value.strip())
    if not match:
        return None
    return int(match.group(1))


def _article_chunks(text: str, expected_article_num: int | None = None) -> tuple[list[str], list[str], int | None]:
    matches = list(ARTICLE_RE.finditer(text))
    if not matches:
        return _fallback_chunk(text), [], expected_article_num

    chunks: list[str] = []
    article_nums: list[str] = []
    accepted_matches: list[re.Match[str]] = []
    next_expected = expected_article_num

    for match in matches:
        article_no = match.group(1)
        article_int = _article_number_as_int(article_no)
        if next_expected is not None:
            if article_int != next_expected:
                continue
            next_expected += 1
        accepted_matches.append(match)

    if not accepted_matches:
        return _fallback_chunk(text), [], expected_article_num

    prefix = text[: accepted_matches[0].start()].strip()
    if prefix:
        for chunk in _fallback_chunk(prefix):
            chunks.append(chunk)
            article_nums.append("")

    for idx, match in enumerate(accepted_matches):
        article_start = match.start()
        article_end = accepted_matches[idx + 1].start() if idx + 1 < len(accepted_matches) else len(text)
        article_text = text[article_start:article_end].strip()
        article_no = match.group(1)

        if not article_text:
            continue

        if len(article_text) <= 2600:
            chunks.append(article_text)
            article_nums.append(article_no)
            continue

        sub_chunks = _fallback_chunk(article_text, chunk_size=2200, overlap=220)
        chunks.extend(sub_chunks)
        article_nums.extend([article_no] * len(sub_chunks))

    return chunks, article_nums, next_expected


def _normalize_pdf_text(text: str) -> str:
    text = _SPACED_ARTICLE_RE.sub("Article", text)
    text = _SPACED_CHAPTER_RE.sub("Chapter", text)
    text = _SPACED_ANNEX_RE.sub("Annex", text)
    text = _INLINE_PDF_ARTICLE_RE.sub(lambda match: f"\n{match.group(1).strip()}", text)
    return text


def _process_pdf(text: str) -> dict[str, Any]:
    pages = _split_pages(text)
    if not pages:
        return {
            "legal_chunks": [],
            "corpus": "",
            "doc_summary": "",
            "page_row_num": 1,
            "chapter_num": 0,
            "article_num": "NA",
            "annex_num": "NA",
        }

    current_chapter = 0
    current_article = "NA"
    current_annex = "NA"
    corpus_segments: list[str] = []
    all_chunks: list[str] = []

    first_chapter = 0
    first_article = "NA"
    first_annex = "NA"
    expected_article_num: int | None = 1

    for page_index, page_text in enumerate(pages, start=1):
        page_text = _normalize_pdf_text(page_text)
        for match in CHAPTER_RE.finditer(page_text):
            current_chapter = _roman_to_int(match.group(1), default=current_chapter)
            if first_chapter == 0:
                first_chapter = current_chapter

        for match in ANNEX_RE.finditer(page_text):
            current_annex = match.group(1).upper()
            if first_annex == "NA":
                first_annex = current_annex

        page_chunks, page_articles, expected_article_num = _article_chunks(page_text, expected_article_num=expected_article_num)
        if not page_chunks:
            page_chunks = _fallback_chunk(page_text)
            page_articles = [current_article] * len(page_chunks)

        for idx, chunk in enumerate(page_chunks):
            article_from_chunk = page_articles[idx] if idx < len(page_articles) else current_article
            if article_from_chunk and article_from_chunk != "NA":
                current_article = article_from_chunk
                if first_article == "NA":
                    first_article = current_article

            annotated_chunk = _annotate_chunk(page_index, current_chapter, current_article, current_annex, chunk)
            all_chunks.append(annotated_chunk)
            corpus_segments.append(annotated_chunk)

    full_summary = " ".join(segment for segment in pages).strip()
    if len(full_summary) > 1800:
        full_summary = full_summary[:1800].rstrip()

    return {
        "legal_chunks": all_chunks,
        "corpus": "\n\n".join(corpus_segments),
        "doc_summary": full_summary,
        "page_row_num": 1,
        "chapter_num": first_chapter,
        "article_num": first_article,
        "annex_num": first_annex,
    }


def decode_file_data(file_data: Any) -> bytes:
    if isinstance(file_data, dict):
        raw = file_data.get("data")
    else:
        raw = file_data

    if not raw:
        return b""

    if not isinstance(raw, str):
        return b""

    encoded = raw.strip()
    if "," in encoded and encoded.lower().startswith("data:"):
        encoded = encoded.split(",", 1)[1]

    try:
        return base64.b64decode(encoded)
    except Exception:
        return b""


def _process_xlsx(file_bytes: bytes, fallback_text: str = "") -> dict[str, Any]:
    if not file_bytes:
        return {
            "legal_chunks": [fallback_text] if fallback_text.strip() else [],
            "corpus": fallback_text,
            "doc_summary": "",
            "page_row_num": 1,
            "chapter_num": 0,
            "article_num": "NA",
            "annex_num": "NA",
        }

    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    legal_chunks: list[str] = []
    corpus_rows: list[str] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_row = rows[0]
        headers = [str(cell).strip() if cell is not None else f"column_{idx + 1}" for idx, cell in enumerate(header_row)]

        for row_idx, row_values in enumerate(rows[1:], start=2):
            if all(value is None or str(value).strip() == "" for value in row_values):
                continue

            row_pairs: list[str] = []
            max_len = max(len(headers), len(row_values))
            for col_idx in range(max_len):
                header = headers[col_idx] if col_idx < len(headers) and headers[col_idx] else f"column_{col_idx + 1}"
                value = row_values[col_idx] if col_idx < len(row_values) else ""
                value_str = "" if value is None else str(value).strip()
                row_pairs.append(f"{header}: {value_str}")

            row_text = " | ".join(row_pairs)
            chunk = f"[page_row_num={row_idx}] {row_text}"
            legal_chunks.append(chunk)
            corpus_rows.append(row_text)

    workbook.close()

    corpus = "\n".join(corpus_rows)
    return {
        "legal_chunks": legal_chunks,
        "corpus": corpus,
        "doc_summary": "",
        "page_row_num": 2,
        "chapter_num": 0,
        "article_num": "NA",
        "annex_num": "NA",
    }


def _html_to_text(raw_html: str) -> str:
    cleaned = _HTML_SCRIPT_STYLE_RE.sub(" ", raw_html)
    cleaned = _HTML_TAG_RE.sub(" ", cleaned)
    cleaned = html.unescape(cleaned)
    cleaned = cleaned.replace("\r", "\n")
    cleaned = _HTML_SPACE_RE.sub(" ", cleaned)
    cleaned = _HTML_NEWLINE_RE.sub("\n\n", cleaned)
    return _clean_text(cleaned)


def _process_html(text: str) -> dict[str, Any]:
    parser = EurLexHtmlParser()
    try:
        parser.feed(text)
        parser.close()
    except Exception:
        parser.blocks = []

    if parser.blocks:
        legal_chunks: list[str] = []
        corpus_segments: list[str] = []
        for block in parser.blocks:
            article_num = block.number if block.kind == "article" else f"recital_{block.number}"
            annex_num = "NA"
            for body in _prefix_split_chunks(block.text):
                annotated = _annotate_chunk(
                    page_row_num=len(legal_chunks) + 1,
                    chapter_num=block.chapter_num,
                    article_num=article_num,
                    annex_num=annex_num,
                    body=body,
                )
                legal_chunks.append(annotated)
                corpus_segments.append(annotated)

        summary = "\n".join(block.text.splitlines()[0] for block in parser.blocks[:20])
        return {
            "legal_chunks": legal_chunks,
            "corpus": "\n\n".join(corpus_segments),
            "doc_summary": summary[:1800].strip(),
            "page_row_num": 1,
            "chapter_num": parser.blocks[0].chapter_num,
            "article_num": parser.blocks[0].number if parser.blocks[0].kind == "article" else f"recital_{parser.blocks[0].number}",
            "annex_num": "NA",
        }

    plain_text = _html_to_text(text)
    chunks = _fallback_chunk(plain_text, chunk_size=2200, overlap=220)
    summary = plain_text[:1800].strip()
    return {
        "legal_chunks": chunks,
        "corpus": plain_text,
        "doc_summary": summary,
        "page_row_num": 1,
        "chapter_num": 0,
        "article_num": "NA",
        "annex_num": "NA",
    }


def chunk_single_document(data: dict[str, Any]) -> dict[str, Any]:
    text = str(data.get("text") or "")
    doc_name = str(data.get("doc_name") or "")
    source_path = str(data.get("source_path") or "")
    file_data = data.get("file_data")

    is_pdf = doc_name.lower().endswith(".pdf") or source_path.lower().endswith(".pdf")
    is_xlsx = (
        doc_name.lower().endswith(".xlsx")
        or source_path.lower().endswith(".xlsx")
        or doc_name.lower().endswith(".xlsm")
        or source_path.lower().endswith(".xlsm")
    )
    is_html = (
        doc_name.lower().endswith(".html")
        or source_path.lower().endswith(".html")
        or doc_name.lower().endswith(".htm")
        or source_path.lower().endswith(".htm")
        or doc_name.lower().endswith(".xml")
        or source_path.lower().endswith(".xml")
    )

    if not text.strip() and not is_xlsx:
        return {
            "legal_chunks": [],
            "corpus": "",
            "doc_summary": "",
            "page_row_num": 1,
            "chapter_num": 0,
            "article_num": "NA",
            "annex_num": "NA",
            "_warning": "Input text is empty.",
        }

    if is_pdf:
        return _process_pdf(text)

    if is_xlsx:
        file_bytes = decode_file_data(file_data)
        return _process_xlsx(file_bytes=file_bytes, fallback_text=text)

    if is_html:
        return _process_html(text)

    non_pdf_summary = text[:1200].strip()
    return {
        "legal_chunks": [text],
        "corpus": text,
        "doc_summary": non_pdf_summary,
        "page_row_num": 1,
        "chapter_num": 0,
        "article_num": "NA",
        "annex_num": "NA",
    }


@app.post("/chunk")
def chunk_documents() -> Any:
    payload = request.get_json(silent=True) or {}
    values = payload.get("values") or []

    out: list[dict[str, Any]] = []
    for item in values:
        record_id = str(item.get("recordId") or "")
        data = item.get("data") or {}
        parsed = chunk_single_document(data)
        warning = parsed.pop("_warning", None)

        item_out: dict[str, Any] = {"recordId": record_id, "data": parsed}
        if warning:
            item_out["warnings"] = [{"message": warning}]
        out.append(item_out)

    return jsonify({"values": out})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)
