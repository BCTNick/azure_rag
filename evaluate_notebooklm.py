from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from evaluate_rag import (
    DEFAULT_JUDGE_DEPLOYMENT,
    JudgeClient,
    judge_payload_to_columns,
    pretty_json,
)
from src.config import load_settings


INPUT_PATH = Path("input_data") / "evaluation" / "notebook_lm.xlsx"
INPUT_SHEET = "notebook_lm_answers"
OUTPUT_ROOT = Path("output_data") / "eval_runs"

REQUIRED_COLUMNS = [
    "source",
    "Question ID",
    "Question",
    "Official Answer",
    "NotebookLM Answer",
    "NotebookLM Citations",
]

NOTEBOOKLM_OUTPUT_COLUMNS = [
    "NotebookLM Answer Overall Score",
    "NotebookLM Correctness Score",
    "NotebookLM Faithfulness Score",
    "NotebookLM Answer Relevance Score",
    "NotebookLM Completeness Score",
    "NotebookLM Citation Precision Score",
    "NotebookLM Citation Recall Score",
    "NotebookLM Unsupported Claims Score",
    "NotebookLM Answer Verdict",
    "NotebookLM Answer Rationale",
    "NotebookLM Answer Missing Points",
    "NotebookLM Answer Unsupported Claims",
    "NotebookLM Answer Raw JSON",
    "NotebookLM Evaluation Error",
]

JUDGE_TO_NOTEBOOKLM_COLUMNS = {
    "Answer Overall Score": "NotebookLM Answer Overall Score",
    "Answer Correctness Score": "NotebookLM Correctness Score",
    "Faithfulness Score": "NotebookLM Faithfulness Score",
    "Answer Relevance Score": "NotebookLM Answer Relevance Score",
    "Completeness Score": "NotebookLM Completeness Score",
    "Citation Precision Score": "NotebookLM Citation Precision Score",
    "Citation Recall Score": "NotebookLM Citation Recall Score",
    "Unsupported Claims Score": "NotebookLM Unsupported Claims Score",
    "Answer Verdict": "NotebookLM Answer Verdict",
    "Answer Rationale": "NotebookLM Answer Rationale",
    "Answer Missing Points": "NotebookLM Answer Missing Points",
    "Answer Unsupported Claims": "NotebookLM Answer Unsupported Claims",
    "Answer Raw JSON": "NotebookLM Answer Raw JSON",
    "Evaluation Error": "NotebookLM Evaluation Error",
}


def main() -> None:
    args = parse_args()
    load_dotenv()

    input_path = Path(args.input)
    if args.in_place:
        output_path = input_path
    else:
        run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        output_dir = OUTPUT_ROOT / run_id
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"notebooklm_answer_evaluated_{run_id}.xlsx"

    workbook = load_workbook(input_path)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(f"NotebookLM sheet not found: {args.sheet}")

    sheet = workbook[args.sheet]
    headers = prepare_output_sheet(sheet)
    workbook.save(output_path)

    judge_deployment = os.getenv("AZURE_OPENAI_JUDGE_DEPLOYMENT", DEFAULT_JUDGE_DEPLOYMENT).strip() or DEFAULT_JUDGE_DEPLOYMENT
    judge_client: JudgeClient | None = None
    if not args.dry_run:
        settings = load_settings()
        judge_client = JudgeClient(settings.azure_openai_endpoint, settings.azure_openai_api_key)

    row_numbers = list(range(args.start_row, sheet.max_row + 1))
    if args.question_id:
        row_numbers = [
            row_num
            for row_num in row_numbers
            if str(sheet.cell(row=row_num, column=headers["Question ID"]).value or "").strip() == args.question_id
        ]

    print(f"Input: {input_path}")
    print(f"Output: {output_path}")
    print(f"Sheet: {args.sheet}")
    print(f"Rows available: {len(row_numbers)}")
    print(f"Judge deployment: {judge_deployment}")
    print("Judge method: evaluate_rag.JudgeClient.evaluate with the same answer rubric and high reasoning effort.")

    preflight = summarize_selected_rows(sheet, headers, row_numbers)
    print(
        "Preflight: "
        f"answered={len(preflight['answered'])}, "
        f"missing={len(preflight['missing'])}, "
        f"generation_failures={len(preflight['generation_failures'])}, "
        f"already_scored={len(preflight['already_scored'])}"
    )
    if preflight["missing"]:
        print("Missing NotebookLM answers:")
        for row_num, question_id in preflight["missing"]:
            print(f"  row {row_num}: {question_id}")
    if preflight["generation_failures"]:
        print("NotebookLM generation-failure rows will be judged as collected answers unless you remove or skip them:")
        for row_num, question_id in preflight["generation_failures"]:
            print(f"  row {row_num}: {question_id}")
    if preflight["metadata_prompt_rows"]:
        print(
            "Warning: NotebookLM raw responses show that benchmark metadata "
            "(for example regulation/topic/article/template) was included in the collection prompt. "
            "The RAG answer run uses only the Question column."
        )
    if args.require_complete and preflight["missing"]:
        raise ValueError("Selected rows are not complete; rerun without --require-complete or collect the missing answers first.")

    judged = 0
    skipped = 0
    failed = 0

    for row_num in row_numbers:
        if args.limit is not None and judged >= args.limit:
            break

        row_values = row_to_dict(sheet, headers, row_num)
        question_id = str(row_values.get("Question ID") or f"row-{row_num}").strip()
        question = str(row_values.get("Question") or "").strip()
        official_answer = str(row_values.get("Official Answer") or "").strip()
        notebooklm_answer = str(row_values.get("NotebookLM Answer") or "").strip()
        notebooklm_citations_text = str(row_values.get("NotebookLM Citations") or "").strip()
        existing_score = row_values.get("NotebookLM Answer Overall Score")

        if existing_score not in (None, "") and not args.overwrite:
            skipped += 1
            continue

        if not question:
            write_outputs(sheet, headers, row_num, {"NotebookLM Evaluation Error": "Missing benchmark question."})
            skipped += 1
            workbook.save(output_path)
            continue

        if not official_answer:
            write_outputs(sheet, headers, row_num, {"NotebookLM Evaluation Error": "Missing official benchmark answer."})
            skipped += 1
            workbook.save(output_path)
            continue

        if not notebooklm_answer:
            write_outputs(sheet, headers, row_num, {"NotebookLM Evaluation Error": "Missing NotebookLM answer; row not judged."})
            skipped += 1
            workbook.save(output_path)
            continue

        print(f"[{judged + 1}] {question_id}")

        if args.dry_run:
            parsed_citations = parse_notebooklm_citations(notebooklm_citations_text)
            print(f"  dry run: answer_chars={len(notebooklm_answer)} citations={len(parsed_citations)}")
            judged += 1
            continue

        try:
            parsed_citations = parse_notebooklm_citations(notebooklm_citations_text)
            if judge_client is None:
                raise RuntimeError("Judge client was not initialized.")
            judge_payload = judge_client.evaluate(
                deployment=judge_deployment,
                question=question,
                official_answer=official_answer,
                rag_answer=notebooklm_answer,
                rag_citations=parsed_citations,
                retrieved_context=build_notebooklm_context(notebooklm_citations_text),
                source=str(row_values.get("source") or ""),
                question_id=question_id,
            )
            write_outputs(sheet, headers, row_num, notebooklm_payload_to_columns(judge_payload))
            judged += 1
        except Exception as ex:
            failed += 1
            write_outputs(sheet, headers, row_num, {"NotebookLM Evaluation Error": f"Judge failed: {ex}"})

        workbook.save(output_path)

    workbook.save(output_path)
    workbook.close()

    print(f"Judged: {judged}")
    print(f"Skipped: {skipped}")
    print(f"Failed: {failed}")
    print(f"Saved evaluated workbook: {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate NotebookLM answers with the same answer-level judge metrics used for the RAG system."
    )
    parser.add_argument("--input", default=str(INPUT_PATH), help="NotebookLM workbook to evaluate.")
    parser.add_argument("--sheet", default=INPUT_SHEET, help="Worksheet containing NotebookLM answers.")
    parser.add_argument("--limit", type=int, default=None, help="Judge at most N eligible rows.")
    parser.add_argument("--start-row", type=int, default=2, help="First worksheet row to consider.")
    parser.add_argument("--question-id", default=None, help="Judge only one Question ID.")
    parser.add_argument("--overwrite", action="store_true", help="Rejudge rows that already have NotebookLM scores.")
    parser.add_argument("--in-place", action="store_true", help="Write scores back to the input workbook instead of a timestamped copy.")
    parser.add_argument("--dry-run", action="store_true", help="Validate row selection and citation parsing without calling the judge model.")
    parser.add_argument("--require-complete", action="store_true", help="Stop if any selected row has no NotebookLM answer.")
    return parser.parse_args()


def prepare_output_sheet(sheet: Any) -> dict[str, int]:
    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(sheet[1], start=1) if cell.value}
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"NotebookLM workbook is missing required columns: {', '.join(missing)}")

    for column in NOTEBOOKLM_OUTPUT_COLUMNS:
        if column not in headers:
            col_idx = sheet.max_column + 1
            sheet.cell(row=1, column=col_idx).value = column
            headers[column] = col_idx

    format_output_sheet(sheet, headers)
    return headers


def format_output_sheet(sheet: Any, headers: dict[str, int]) -> None:
    notebook_fill = PatternFill("solid", fgColor="806000")
    judge_fill = PatternFill("solid", fgColor="7030A0")
    header_font = Font(color="FFFFFF", bold=True)

    for column_name in NOTEBOOKLM_OUTPUT_COLUMNS:
        col_idx = headers[column_name]
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = judge_fill if "Score" in column_name or "Raw JSON" in column_name else notebook_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    width_by_header = {
        "NotebookLM Answer": 70,
        "NotebookLM Citations": 55,
        "NotebookLM Answer Rationale": 65,
        "NotebookLM Answer Missing Points": 55,
        "NotebookLM Answer Unsupported Claims": 55,
        "NotebookLM Answer Raw JSON": 55,
        "NotebookLM Evaluation Error": 45,
    }
    for column_name, width in width_by_header.items():
        if column_name in headers:
            sheet.column_dimensions[sheet.cell(row=1, column=headers[column_name]).column_letter].width = width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def row_to_dict(sheet: Any, headers: dict[str, int], row_num: int) -> dict[str, Any]:
    return {header: sheet.cell(row=row_num, column=col_idx).value for header, col_idx in headers.items()}


def write_outputs(sheet: Any, headers: dict[str, int], row_num: int, values: dict[str, Any]) -> None:
    for column, value in values.items():
        if column not in headers:
            raise ValueError(f"Unknown output column: {column}")
        sheet.cell(row=row_num, column=headers[column]).value = value


def summarize_selected_rows(sheet: Any, headers: dict[str, int], row_numbers: list[int]) -> dict[str, list[tuple[int, str]]]:
    summary: dict[str, list[tuple[int, str]]] = {
        "answered": [],
        "missing": [],
        "generation_failures": [],
        "already_scored": [],
        "metadata_prompt_rows": [],
    }
    for row_num in row_numbers:
        question_id = str(sheet.cell(row=row_num, column=headers["Question ID"]).value or f"row-{row_num}").strip()
        answer = str(sheet.cell(row=row_num, column=headers["NotebookLM Answer"]).value or "").strip()
        raw_col = headers.get("NotebookLM Raw Response")
        raw = str(sheet.cell(row=row_num, column=raw_col).value or "").strip() if raw_col else ""
        score = sheet.cell(row=row_num, column=headers["NotebookLM Answer Overall Score"]).value

        if answer:
            summary["answered"].append((row_num, question_id))
        else:
            summary["missing"].append((row_num, question_id))

        if "answer-generation failure" in answer.lower() or "difficolt" in raw.lower():
            summary["generation_failures"].append((row_num, question_id))

        if score not in (None, ""):
            summary["already_scored"].append((row_num, question_id))

        if raw and any(marker in raw for marker in ("Regulation reference:", "Topic:", "ARTICLE:", "TEMPLATE:")):
            summary["metadata_prompt_rows"].append((row_num, question_id))

    return summary


def notebooklm_payload_to_columns(payload: dict[str, Any]) -> dict[str, Any]:
    rag_columns = judge_payload_to_columns(payload)
    return {notebooklm_column: rag_columns[rag_column] for rag_column, notebooklm_column in JUDGE_TO_NOTEBOOKLM_COLUMNS.items()}


def parse_notebooklm_citations(citations_text: str) -> list[dict[str, Any]]:
    text = citations_text.strip()
    if not text or re.search(r"^not provided by notebooklm\.?$", text, flags=re.IGNORECASE):
        return []

    citations: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    pending_rank: int | None = None

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        rank_match = re.match(r"^\[?(\d+)\]?[.)]?$", line)
        if rank_match:
            pending_rank = int(rank_match.group(1))
            continue

        starts_new_citation = "doc=" in line or re.match(r"^\[?\d+\]?[.)]\s+", line)
        if starts_new_citation:
            if current:
                citations.append(current)
            rank, body = split_inline_rank(line)
            current = {
                "rank": rank or pending_rank or len(citations) + 1,
                "raw": body,
                **extract_citation_fields(body),
            }
            pending_rank = None
        elif current:
            current["raw"] = f"{current['raw']}\n{line}".strip()
            current["support"] = line if "support" not in current else f"{current['support']}\n{line}"
        else:
            current = {"rank": pending_rank or len(citations) + 1, "raw": line, "support": line}
            pending_rank = None

    if current:
        citations.append(current)

    return citations or [{"rank": 1, "raw": text}]


def split_inline_rank(line: str) -> tuple[int | None, str]:
    match = re.match(r"^\[?(\d+)\]?[.)]\s+(.*)$", line)
    if not match:
        return None, line
    return int(match.group(1)), match.group(2).strip()


def extract_citation_fields(line: str) -> dict[str, str]:
    keys = ["doc", "type", "page_row", "chapter", "article", "annex", "template", "field"]
    matches = list(re.finditer(r"\b(" + "|".join(keys) + r")=", line))
    fields: dict[str, str] = {}
    if not matches:
        return fields

    for index, match in enumerate(matches):
        key = match.group(1)
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(line)
        fields[key] = line[start:end].strip(" -")

    return fields


def build_notebooklm_context(citations_text: str) -> str:
    if citations_text.strip():
        return (
            "NotebookLM does not expose the full retrieved context through the web interface. "
            "The evaluator therefore receives the citations and source-support notes produced by NotebookLM:\n"
            f"{citations_text.strip()}"
        )
    return "NotebookLM did not provide citations or retrieved context for this answer."


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted by user.")
        sys.exit(1)
    except Exception as ex:
        print(f"ERROR: {ex}")
        sys.exit(1)
