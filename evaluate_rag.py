from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.chat import ChatQueryEngine
from src.config import load_settings


BENCHMARK_PATH = Path("input_data") / "evaluation" / "dora_qas_eiopa_eba_benchmark.xlsx"
BENCHMARK_SHEET = "benchmark_qas"
OUTPUT_ROOT = Path("output_data") / "eval_runs"
DEFAULT_JUDGE_DEPLOYMENT = "gpt-5.4-mini-2"

REQUIRED_COLUMNS = ["source", "Question ID", "Question", "Official Answer"]
OUTPUT_COLUMNS = [
    "RAG Answer",
    "RAG Citations JSON",
    "RAG Trace ID",
    "RAG Active Query",
    "RAG Profile JSON",
    "RAG Retrieval Skipped",
    "Answer Overall Score",
    "Answer Correctness Score",
    "Faithfulness Score",
    "Answer Relevance Score",
    "Completeness Score",
    "Citation Precision Score",
    "Citation Recall Score",
    "Unsupported Claims Score",
    "Answer Verdict",
    "Answer Rationale",
    "Answer Missing Points",
    "Answer Unsupported Claims",
    "Answer Raw JSON",
    "Retrieval Reference Evaluation Needed",
    "Retrieval Gold References JSON",
    "Retrieval Retrieved References JSON",
    "Retrieval Chunk Judgments JSON",
    "Retrieval Judged Chunk Count",
    "Retrieval Precision@5",
    "Retrieval Precision@10",
    "Retrieval Precision@ActualK",
    "Retrieval MRR",
    "Retrieval nDCG@5",
    "Retrieval nDCG@10",
    "Retrieval Hit@5",
    "Retrieval Hit@10",
    "Retrieval Verdict",
    "Retrieval Rationale",
    "Retrieval Raw JSON",
    "Evaluation Error",
]

JUDGE_JSON_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "overall_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "answer_correctness_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "faithfulness_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "answer_relevance_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "completeness_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "citation_precision_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "citation_recall_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "unsupported_claims_score": {"type": "integer", "minimum": 0, "maximum": 5},
        "verdict": {"type": "string"},
        "rationale": {"type": "string"},
        "missing_points": {"type": "string"},
        "unsupported_claims": {"type": "string"},
    },
    "required": [
        "overall_score",
        "answer_correctness_score",
        "faithfulness_score",
        "answer_relevance_score",
        "completeness_score",
        "citation_precision_score",
        "citation_recall_score",
        "unsupported_claims_score",
        "verdict",
        "rationale",
        "missing_points",
        "unsupported_claims",
    ],
}

RETRIEVAL_JUDGE_JSON_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "gold_references": {
            "type": "array",
            "items": {"type": "string"},
        },
        "retrieved_references": {
            "type": "array",
            "items": {"type": "string"},
        },
        "chunk_judgments": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "rank": {"type": "integer", "minimum": 1},
                    "relevance_grade": {"type": "integer", "minimum": 0, "maximum": 3},
                    "is_relevant": {"type": "boolean"},
                    "matched_references": {
                        "type": "array",
                        "items": {"type": "string"},
                    },
                    "rationale": {"type": "string"},
                },
                "required": ["rank", "relevance_grade", "is_relevant", "matched_references", "rationale"],
            },
        },
        "verdict": {"type": "string"},
        "rationale": {"type": "string"},
    },
    "required": [
        "gold_references",
        "retrieved_references",
        "chunk_judgments",
        "verdict",
        "rationale",
    ],
}


def main() -> None:
    args = parse_args()
    load_dotenv()

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    output_dir = OUTPUT_ROOT / run_id
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"dora_qas_eiopa_eba_evaluated_{run_id}.xlsx"

    workbook = load_workbook(BENCHMARK_PATH)
    if BENCHMARK_SHEET not in workbook.sheetnames:
        raise ValueError(f"Benchmark sheet not found: {BENCHMARK_SHEET}")

    sheet = workbook[BENCHMARK_SHEET]
    headers = prepare_output_sheet(sheet)
    workbook.save(output_path)

    settings = load_settings()
    chat_engine = ChatQueryEngine(settings, enable_trace_log=True)
    judge_client = JudgeClient(settings.azure_openai_endpoint, settings.azure_openai_api_key)
    judge_deployment = os.getenv("AZURE_OPENAI_JUDGE_DEPLOYMENT", DEFAULT_JUDGE_DEPLOYMENT).strip() or DEFAULT_JUDGE_DEPLOYMENT

    row_numbers = list(range(2, sheet.max_row + 1))
    if args.limit is not None:
        row_numbers = row_numbers[: args.limit]

    print(f"Evaluation run: {run_id}")
    print(f"Input: {BENCHMARK_PATH}")
    print(f"Output: {output_path}")
    print(f"Rows to evaluate: {len(row_numbers)}")
    print(f"Judge enabled: {not args.no_judge}; deployment: {judge_deployment}")

    for position, row_num in enumerate(row_numbers, start=1):
        row_values = row_to_dict(sheet, headers, row_num)
        question_id = str(row_values.get("Question ID") or f"row-{row_num}")
        question = str(row_values.get("Question") or "").strip()
        official_answer = str(row_values.get("Official Answer") or "").strip()
        errors: list[str] = []

        print(f"[{position}/{len(row_numbers)}] {question_id}")

        if not question:
            write_outputs(sheet, headers, row_num, {"Evaluation Error": "Missing benchmark question."})
            workbook.save(output_path)
            continue

        try:
            rag_result = chat_engine.answer_once(question)
            rag_answer = str(rag_result.get("answer") or "")
            rag_citations = rag_result.get("citations") or []
            rag_context = str(rag_result.get("retrieved_context") or "")
            write_outputs(
                sheet,
                headers,
                row_num,
                {
                    "RAG Answer": rag_answer,
                    "RAG Citations JSON": pretty_json(rag_citations),
                    "RAG Trace ID": rag_result.get("trace_id", ""),
                    "RAG Active Query": rag_result.get("active_query", ""),
                    "RAG Profile JSON": pretty_json(rag_result.get("profile", {})),
                    "RAG Retrieval Skipped": bool(rag_result.get("retrieval_skipped", False)),
                },
            )
        except Exception as ex:
            errors.append(f"RAG failed: {ex}")
            write_outputs(sheet, headers, row_num, {"Evaluation Error": "\n".join(errors)})
            workbook.save(output_path)
            continue

        if not args.no_judge:
            try:
                judge_payload = judge_client.evaluate(
                    deployment=judge_deployment,
                    question=question,
                    official_answer=official_answer,
                    rag_answer=rag_answer,
                    rag_citations=rag_citations,
                    retrieved_context=rag_context,
                    source=str(row_values.get("source") or ""),
                    question_id=question_id,
                )
                write_outputs(sheet, headers, row_num, judge_payload_to_columns(judge_payload))
            except Exception as ex:
                errors.append(f"Judge failed: {ex}")

            if should_evaluate_retrieval_references(row_values):
                try:
                    retrieval_payload = judge_client.evaluate_retrieval_references(
                        deployment=judge_deployment,
                        question=question,
                        official_answer=official_answer,
                        benchmark_article=str(row_values.get("Article") or ""),
                        benchmark_template=str(row_values.get("Template") or ""),
                        regulation_reference=str(row_values.get("Regulation Reference") or ""),
                        qa_topic=str(row_values.get("QA Topic") or ""),
                        rag_citations=rag_citations,
                        retrieved_context=rag_context,
                        question_id=question_id,
                    )
                    write_outputs(sheet, headers, row_num, retrieval_payload_to_columns(retrieval_payload, needed=True))
                except Exception as ex:
                    errors.append(f"Retrieval reference judge failed: {ex}")
            else:
                write_outputs(
                    sheet,
                    headers,
                    row_num,
                    {
                        "Retrieval Reference Evaluation Needed": False,
                        "Retrieval Verdict": "Skipped: benchmark Article and Template are blank or not meaningful.",
                    },
                )

        if errors:
            write_outputs(sheet, headers, row_num, {"Evaluation Error": "\n".join(errors)})
        workbook.save(output_path)

    workbook.save(output_path)
    workbook.close()
    print(f"Saved evaluated workbook: {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate the legal RAG against the DORA EIOPA/EBA benchmark workbook.")
    parser.add_argument("--limit", type=int, default=None, help="Evaluate only the first N benchmark rows.")
    parser.add_argument("--no-judge", action="store_true", help="Collect RAG answers only; skip judge model calls.")
    return parser.parse_args()


def prepare_output_sheet(sheet: Any) -> dict[str, int]:
    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(sheet[1], start=1) if cell.value}
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Benchmark workbook is missing required columns: {', '.join(missing)}")

    for column in OUTPUT_COLUMNS:
        if column not in headers:
            col_idx = sheet.max_column + 1
            sheet.cell(row=1, column=col_idx).value = column
            headers[column] = col_idx

    format_output_sheet(sheet, headers)
    return headers


def format_output_sheet(sheet: Any, headers: dict[str, int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    judge_fill = PatternFill("solid", fgColor="7030A0")
    rag_fill = PatternFill("solid", fgColor="548235")
    retrieval_fill = PatternFill("solid", fgColor="C65911")
    header_font = Font(color="FFFFFF", bold=True)

    for column_name, col_idx in headers.items():
        cell = sheet.cell(row=1, column=col_idx)
        if column_name.startswith("Answer") or column_name in {
            "Faithfulness Score",
            "Completeness Score",
            "Citation Precision Score",
            "Citation Recall Score",
            "Unsupported Claims Score",
        }:
            cell.fill = judge_fill
        elif column_name.startswith("Retrieval"):
            cell.fill = retrieval_fill
        elif column_name.startswith("RAG") or column_name == "Evaluation Error":
            cell.fill = rag_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    width_by_header = {
        "RAG Answer": 70,
        "RAG Citations JSON": 45,
        "RAG Active Query": 45,
        "RAG Profile JSON": 35,
        "Answer Rationale": 65,
        "Answer Missing Points": 55,
        "Answer Unsupported Claims": 55,
        "Answer Raw JSON": 55,
        "Retrieval Gold References JSON": 45,
        "Retrieval Retrieved References JSON": 45,
        "Retrieval Chunk Judgments JSON": 55,
        "Retrieval Rationale": 65,
        "Retrieval Raw JSON": 55,
        "Evaluation Error": 45,
    }
    for column_name, width in width_by_header.items():
        if column_name in headers:
            sheet.column_dimensions[sheet.cell(row=1, column=headers[column_name]).column_letter].width = width

    for row_idx in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 90

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for table in sheet.tables.values():
        table.ref = sheet.dimensions


def row_to_dict(sheet: Any, headers: dict[str, int], row_num: int) -> dict[str, Any]:
    return {header: sheet.cell(row=row_num, column=col_idx).value for header, col_idx in headers.items()}


def write_outputs(sheet: Any, headers: dict[str, int], row_num: int, values: dict[str, Any]) -> None:
    for column, value in values.items():
        if column not in headers:
            raise ValueError(f"Unknown output column: {column}")
        sheet.cell(row=row_num, column=headers[column]).value = value


def pretty_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def is_meaningful_reference(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text not in {"", "n/a", "na", "none", "null", "not applicable", "-"}


def should_evaluate_retrieval_references(row_values: dict[str, Any]) -> bool:
    return is_meaningful_reference(row_values.get("Article")) or is_meaningful_reference(row_values.get("Template"))


def judge_payload_to_columns(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "Answer Overall Score": int(payload.get("overall_score", 0)),
        "Answer Correctness Score": int(payload.get("answer_correctness_score", 0)),
        "Faithfulness Score": int(payload.get("faithfulness_score", 0)),
        "Answer Relevance Score": int(payload.get("answer_relevance_score", 0)),
        "Completeness Score": int(payload.get("completeness_score", 0)),
        "Citation Precision Score": int(payload.get("citation_precision_score", 0)),
        "Citation Recall Score": int(payload.get("citation_recall_score", 0)),
        "Unsupported Claims Score": int(payload.get("unsupported_claims_score", 0)),
        "Answer Verdict": str(payload.get("verdict", "")),
        "Answer Rationale": str(payload.get("rationale", "")),
        "Answer Missing Points": str(payload.get("missing_points", "")),
        "Answer Unsupported Claims": str(payload.get("unsupported_claims", "")),
        "Answer Raw JSON": pretty_json(payload),
        "Evaluation Error": "",
    }


def retrieval_payload_to_columns(payload: dict[str, Any], needed: bool) -> dict[str, Any]:
    metrics = compute_retrieval_metrics(payload.get("chunk_judgments", []))
    return {
        "Retrieval Reference Evaluation Needed": needed,
        "Retrieval Gold References JSON": pretty_json(payload.get("gold_references", [])),
        "Retrieval Retrieved References JSON": pretty_json(payload.get("retrieved_references", [])),
        "Retrieval Chunk Judgments JSON": pretty_json(payload.get("chunk_judgments", [])),
        "Retrieval Judged Chunk Count": metrics["judged_chunk_count"],
        "Retrieval Precision@5": metrics["precision_at_5"],
        "Retrieval Precision@10": metrics["precision_at_10"],
        "Retrieval Precision@ActualK": metrics["precision_at_actual_k"],
        "Retrieval MRR": metrics["mrr"],
        "Retrieval nDCG@5": metrics["ndcg_at_5"],
        "Retrieval nDCG@10": metrics["ndcg_at_10"],
        "Retrieval Hit@5": metrics["hit_at_5"],
        "Retrieval Hit@10": metrics["hit_at_10"],
        "Retrieval Verdict": str(payload.get("verdict", "")),
        "Retrieval Rationale": str(payload.get("rationale", "")),
        "Retrieval Raw JSON": pretty_json(payload),
        "Evaluation Error": "",
    }


def compute_retrieval_metrics(chunk_judgments: Any) -> dict[str, float | int]:
    judgments = normalize_chunk_judgments(chunk_judgments)
    grades_by_rank = {judgment["rank"]: judgment["relevance_grade"] for judgment in judgments}
    judged_count = max(grades_by_rank) if grades_by_rank else 0

    return {
        "judged_chunk_count": judged_count,
        "precision_at_5": precision_at_k(grades_by_rank, 5, judged_count),
        "precision_at_10": precision_at_k(grades_by_rank, 10, judged_count),
        "precision_at_actual_k": precision_at_k(grades_by_rank, judged_count, judged_count) if judged_count else None,
        "mrr": mean_reciprocal_rank(grades_by_rank),
        "ndcg_at_5": ndcg_at_k(grades_by_rank, 5),
        "ndcg_at_10": ndcg_at_k(grades_by_rank, 10),
        "hit_at_5": hit_at_k(grades_by_rank, 5),
        "hit_at_10": hit_at_k(grades_by_rank, 10),
    }


def normalize_chunk_judgments(chunk_judgments: Any) -> list[dict[str, int]]:
    if not isinstance(chunk_judgments, list):
        return []

    normalized: list[dict[str, int]] = []
    seen_ranks: set[int] = set()
    for item in chunk_judgments:
        if not isinstance(item, dict):
            continue
        try:
            rank = int(item.get("rank"))
            grade = int(item.get("relevance_grade"))
        except (TypeError, ValueError):
            continue
        if rank < 1 or rank in seen_ranks:
            continue
        grade = max(0, min(3, grade))
        normalized.append({"rank": rank, "relevance_grade": grade})
        seen_ranks.add(rank)

    return sorted(normalized, key=lambda row: row["rank"])


def is_binary_relevant(grade: int) -> bool:
    return grade >= 2


def precision_at_k(grades_by_rank: dict[int, int], k: int, judged_count: int) -> float | None:
    if k <= 0 or judged_count <= 0:
        return None
    if judged_count < k:
        return None
    relevant = sum(1 for rank in range(1, k + 1) if is_binary_relevant(grades_by_rank.get(rank, 0)))
    return round(relevant / k, 4)


def hit_at_k(grades_by_rank: dict[int, int], k: int) -> int:
    return int(any(is_binary_relevant(grades_by_rank.get(rank, 0)) for rank in range(1, k + 1)))


def mean_reciprocal_rank(grades_by_rank: dict[int, int]) -> float:
    for rank in sorted(grades_by_rank):
        if is_binary_relevant(grades_by_rank[rank]):
            return round(1 / rank, 4)
    return 0.0


def ndcg_at_k(grades_by_rank: dict[int, int], k: int) -> float:
    if k <= 0:
        return 0.0

    grades = [grades_by_rank.get(rank, 0) for rank in range(1, k + 1)]
    dcg = discounted_cumulative_gain(grades)
    ideal_grades = sorted(grades, reverse=True)
    ideal_dcg = discounted_cumulative_gain(ideal_grades)
    if ideal_dcg == 0:
        return 0.0
    return round(dcg / ideal_dcg, 4)


def discounted_cumulative_gain(grades: list[int]) -> float:
    total = 0.0
    for zero_based_idx, grade in enumerate(grades):
        rank = zero_based_idx + 1
        gain = (2**grade) - 1
        total += gain / math.log2(rank + 1)
    return total


class JudgeClient:
    def __init__(self, azure_openai_endpoint: str, api_key: str) -> None:
        self.endpoint = azure_openai_endpoint.rstrip("/")
        self.api_key = api_key

    def evaluate(
        self,
        deployment: str,
        question: str,
        official_answer: str,
        rag_answer: str,
        rag_citations: list[dict[str, Any]],
        retrieved_context: str,
        source: str,
        question_id: str,
    ) -> dict[str, Any]:
        response_json = self._post_responses(
            deployment=deployment,
            messages=self._build_messages(
                question=question,
                official_answer=official_answer,
                rag_answer=rag_answer,
                rag_citations=rag_citations,
                retrieved_context=retrieved_context,
                source=source,
                question_id=question_id,
            ),
            strict_json=True,
            schema_name="rag_answer_evaluation",
            schema=JUDGE_JSON_SCHEMA,
        )
        text = extract_response_text(response_json)
        payload = extract_json_object(text)
        validate_judge_payload(payload)
        return payload

    def evaluate_retrieval_references(
        self,
        deployment: str,
        question: str,
        official_answer: str,
        benchmark_article: str,
        benchmark_template: str,
        regulation_reference: str,
        qa_topic: str,
        rag_citations: list[dict[str, Any]],
        retrieved_context: str,
        question_id: str,
    ) -> dict[str, Any]:
        response_json = self._post_responses(
            deployment=deployment,
            messages=self._build_retrieval_reference_messages(
                question=question,
                official_answer=official_answer,
                benchmark_article=benchmark_article,
                benchmark_template=benchmark_template,
                regulation_reference=regulation_reference,
                qa_topic=qa_topic,
                rag_citations=rag_citations,
                retrieved_context=retrieved_context,
                question_id=question_id,
            ),
            strict_json=True,
            schema_name="rag_retrieval_reference_evaluation",
            schema=RETRIEVAL_JUDGE_JSON_SCHEMA,
        )
        text = extract_response_text(response_json)
        payload = extract_json_object(text)
        validate_retrieval_payload(payload)
        return payload

    def _post_responses(
        self,
        deployment: str,
        messages: list[dict[str, Any]],
        strict_json: bool,
        schema_name: str,
        schema: dict[str, Any],
    ) -> dict[str, Any]:
        body: dict[str, Any] = {
            "model": deployment,
            "input": messages,
            "reasoning": {"effort": "high"},
        }
        if strict_json:
            body["text"] = {
                "format": {
                    "type": "json_schema",
                    "name": schema_name,
                    "strict": True,
                    "schema": schema,
                }
            }

        response = requests.post(
            f"{self.endpoint}/responses",
            headers={"Content-Type": "application/json", "api-key": self.api_key},
            json=body,
            timeout=240,
        )
        if response.status_code >= 400 and strict_json:
            return self._post_responses(
                deployment=deployment,
                messages=messages,
                strict_json=False,
                schema_name=schema_name,
                schema=schema,
            )

        try:
            response.raise_for_status()
        except requests.HTTPError as ex:
            detail = (response.text or "").strip()
            raise RuntimeError(f"Judge Responses API failed with status {response.status_code}: {detail}") from ex
        return response.json()

    @staticmethod
    def _build_messages(
        question: str,
        official_answer: str,
        rag_answer: str,
        rag_citations: list[dict[str, Any]],
        retrieved_context: str,
        source: str,
        question_id: str,
    ) -> list[dict[str, Any]]:
        system_prompt = (
            "You are a strict legal RAG evaluator for a thesis benchmark. "
            "Compare the RAG answer against the official supervisory Q&A answer. "
            "Do not reward verbosity. Do not require identical wording. "
            "Reward legal equivalence, material completeness, and support by cited sources. "
            "Return only valid JSON with the required keys. Scores are integers from 0 to 5, where 5 is best."
        )
        user_prompt = f"""
Question ID: {question_id}
Benchmark source: {source}

Question:
{question}

Official Q&A answer:
{official_answer}

RAG answer:
{rag_answer}

RAG citations:
{pretty_json(rag_citations)}

Retrieved context used by the RAG:
{retrieved_context}

Rubric:
- overall_score: overall legal answer quality compared with the official answer.
- answer_correctness_score: whether the answer reaches the legally correct conclusion.
- faithfulness_score: whether answer claims are supported by the retrieved context and citations.
- answer_relevance_score: whether the answer directly addresses the question asked.
- completeness_score: coverage of material conditions, exceptions, definitions, and qualifications.
- citation_precision_score: whether cited sources actually support the claims for which they are used.
- citation_recall_score: whether the answer cites the essential sources needed to verify the answer.
- unsupported_claims_score: absence of unsupported or invented legal claims; 5 means no unsupported claims.

Return JSON with:
overall_score, answer_correctness_score, faithfulness_score, answer_relevance_score,
completeness_score, citation_precision_score, citation_recall_score, unsupported_claims_score,
verdict, rationale, missing_points, unsupported_claims.
"""
        return [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]

    @staticmethod
    def _build_retrieval_reference_messages(
        question: str,
        official_answer: str,
        benchmark_article: str,
        benchmark_template: str,
        regulation_reference: str,
        qa_topic: str,
        rag_citations: list[dict[str, Any]],
        retrieved_context: str,
        question_id: str,
    ) -> list[dict[str, Any]]:
        system_prompt = (
            "You are a strict legal retrieval evaluator for a DORA RAG thesis benchmark. "
            "Your task is to judge the relevance of each retrieved chunk against the expected legal references "
            "from benchmark metadata. The benchmark references may be informal, "
            "for example 'Article 28, Paragraph 9', '28(9)', 'B_02.02.0090', or a reporting template name. "
            "Normalize equivalent formats before comparing. Return only valid JSON with the required keys. "
            "Do not compute precision, MRR, or nDCG yourself; assign only per-chunk relevance grades."
        )
        user_prompt = f"""
Question ID: {question_id}

Benchmark metadata:
- Regulation Reference: {regulation_reference}
- QA Topic: {qa_topic}
- Article: {benchmark_article}
- Template: {benchmark_template}

Question:
{question}

Official Q&A answer:
{official_answer}

RAG citations, in retrieval/context order:
{pretty_json(rag_citations)}

Retrieved context:
{retrieved_context}

Instructions:
1. Extract normalized expected references from the benchmark Article and Template fields. Use the question and
   official answer only to disambiguate messy references, not to invent unrelated gold references.
2. Extract normalized references from the retrieved citations and context.
3. Create one chunk_judgments item for each retrieved citation/chunk rank shown in the RAG citations.
4. Assign each chunk a relevance_grade:
   - 3 = directly answer-bearing evidence for at least one expected Article/Template reference.
   - 2 = necessary supporting evidence for an expected reference, but not sufficient alone.
   - 1 = related background or same broad legal area, but not enough to answer the benchmark reference.
   - 0 = irrelevant to the expected Article/Template references.
5. Set is_relevant to true only when relevance_grade is 2 or 3. Grade 1 is related background but not
   relevant enough for binary Precision/MRR/Hit calculations.
6. Keep ranks exactly aligned with the RAG citations order. Do not skip ranks unless no citation exists for that rank.

Return JSON with:
gold_references, retrieved_references, chunk_judgments, verdict, rationale.
"""
        return [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]


def extract_response_text(response_json: dict[str, Any]) -> str:
    output_text = response_json.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()

    parts: list[str] = []
    for item in response_json.get("output", []) or []:
        for content in item.get("content", []) or []:
            if isinstance(content, dict):
                text = content.get("text") or content.get("output_text")
                if isinstance(text, str):
                    parts.append(text)
    if parts:
        return "\n".join(parts).strip()

    choices = response_json.get("choices", [])
    if choices:
        content = choices[0].get("message", {}).get("content", "")
        if isinstance(content, str):
            return content.strip()

    raise ValueError(f"Could not extract judge text from response: {pretty_json(response_json)[:2000]}")


def extract_json_object(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", cleaned, flags=re.DOTALL | re.IGNORECASE)
    if fenced:
        cleaned = fenced.group(1)
    elif "{" in cleaned and "}" in cleaned:
        cleaned = cleaned[cleaned.find("{") : cleaned.rfind("}") + 1]

    try:
        payload = json.loads(cleaned)
    except json.JSONDecodeError as ex:
        raise ValueError(f"Judge returned invalid JSON: {text[:2000]}") from ex
    if not isinstance(payload, dict):
        raise ValueError(f"Judge JSON must be an object: {text[:2000]}")
    return payload


def validate_judge_payload(payload: dict[str, Any]) -> None:
    required = JUDGE_JSON_SCHEMA["required"]
    missing = [key for key in required if key not in payload]
    if missing:
        raise ValueError(f"Judge JSON missing required keys: {', '.join(missing)}")

    score_keys = [key for key in required if key.endswith("_score")]
    for key in score_keys:
        value = payload[key]
        if not isinstance(value, int) or not 0 <= value <= 5:
            raise ValueError(f"Judge score {key} must be an integer from 0 to 5, got {value!r}")


def validate_retrieval_payload(payload: dict[str, Any]) -> None:
    required = RETRIEVAL_JUDGE_JSON_SCHEMA["required"]
    missing = [key for key in required if key not in payload]
    if missing:
        raise ValueError(f"Retrieval judge JSON missing required keys: {', '.join(missing)}")

    for key in ("gold_references", "retrieved_references"):
        value = payload[key]
        if not isinstance(value, list) or not all(isinstance(item, str) for item in value):
            raise ValueError(f"Retrieval judge field {key} must be a list of strings, got {value!r}")

    chunk_judgments = payload["chunk_judgments"]
    if not isinstance(chunk_judgments, list):
        raise ValueError(f"Retrieval judge field chunk_judgments must be a list, got {chunk_judgments!r}")

    for item in chunk_judgments:
        if not isinstance(item, dict):
            raise ValueError(f"Each retrieval chunk judgment must be an object, got {item!r}")
        rank = item.get("rank")
        grade = item.get("relevance_grade")
        is_relevant = item.get("is_relevant")
        matched_references = item.get("matched_references")
        if not isinstance(rank, int) or rank < 1:
            raise ValueError(f"Retrieval chunk rank must be a positive integer, got {rank!r}")
        if not isinstance(grade, int) or not 0 <= grade <= 3:
            raise ValueError(f"Retrieval relevance_grade must be an integer from 0 to 3, got {grade!r}")
        if not isinstance(is_relevant, bool):
            raise ValueError(f"Retrieval is_relevant must be boolean, got {is_relevant!r}")
        if is_relevant != is_binary_relevant(grade):
            raise ValueError(f"Retrieval is_relevant must equal relevance_grade >= 2 for rank {rank}")
        if not isinstance(matched_references, list) or not all(isinstance(ref, str) for ref in matched_references):
            raise ValueError(f"Retrieval matched_references must be a list of strings, got {matched_references!r}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted by user.")
        sys.exit(1)
    except Exception as ex:
        print(f"ERROR: {ex}")
        sys.exit(1)
