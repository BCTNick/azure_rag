from __future__ import annotations

import argparse
import csv
import json
import math
import re
import warnings
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median, pstdev
from typing import Any

warnings.filterwarnings("ignore", category=RuntimeWarning, module="numpy")
warnings.filterwarnings("ignore", message="Numpy built with MINGW-W64.*")

from openpyxl import load_workbook


DEFAULT_INPUT = Path("output_data/eval_runs/20260623T094922Z/dora_qas_eiopa_eba_evaluated_20260623T094922Z.xlsx")
DEFAULT_SHEET = "benchmark_qas"
DEFAULT_OUTPUT_DIR = Path("output_data/eval_runs/20260623T094922Z/analysis")
DEFAULT_FIGURE_DIR = Path("latex/images/evaluation")

ANSWER_METRICS = [
    ("Answer Overall Score", "Overall"),
    ("Answer Correctness Score", "Correctness"),
    ("Faithfulness Score", "Faithfulness"),
    ("Answer Relevance Score", "Relevance"),
    ("Completeness Score", "Completeness"),
    ("Citation Precision Score", "Citation precision"),
    ("Citation Recall Score", "Citation recall"),
    ("Unsupported Claims Score", "No unsupported claims"),
]

RETRIEVAL_METRICS = [
    ("Retrieval Precision@5", "Precision@5"),
    ("Retrieval Precision@10", "Precision@10"),
    ("Retrieval Precision@ActualK", "Precision@ActualK"),
    ("Retrieval MRR", "MRR"),
    ("Retrieval nDCG@5", "nDCG@5"),
    ("Retrieval nDCG@10", "nDCG@10"),
    ("Retrieval Hit@5", "Hit@5"),
    ("Retrieval Hit@10", "Hit@10"),
]


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    figure_dir = Path(args.figure_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    figure_dir.mkdir(parents=True, exist_ok=True)

    rows = load_rows(input_path, args.sheet)
    enriched_rows = enrich_rows(rows)

    answer_summary = summarize_metrics(enriched_rows, ANSWER_METRICS)
    retrieval_rows = [row for row in enriched_rows if truthy(row.get("Retrieval Reference Evaluation Needed"))]
    retrieval_summary = summarize_metrics(retrieval_rows, RETRIEVAL_METRICS)
    source_summary = summarize_by_group(enriched_rows, "source")
    topic_summary = summarize_by_group(enriched_rows, "topic_normalized")
    answer_distributions = score_distributions(enriched_rows, ANSWER_METRICS)
    retrieval_grade_distribution = chunk_grade_distribution(retrieval_rows)
    first_relevant_distribution = Counter(first_relevant_rank(row) for row in retrieval_rows)
    correlation_summary = compute_correlations(retrieval_rows)

    analysis = {
        "input_workbook": str(input_path),
        "row_count": len(enriched_rows),
        "retrieval_evaluated_count": len(retrieval_rows),
        "error_count": sum(1 for row in enriched_rows if clean_text(row.get("Evaluation Error"))),
        "answer_summary": answer_summary,
        "retrieval_summary": retrieval_summary,
        "source_summary": source_summary,
        "topic_summary": topic_summary,
        "answer_distributions": answer_distributions,
        "retrieval_grade_distribution": dict(sorted(retrieval_grade_distribution.items())),
        "first_relevant_rank_distribution": stringify_counter(first_relevant_distribution),
        "correlations": correlation_summary,
    }

    write_json(output_dir / "evaluation_analysis_summary.json", analysis)
    write_metric_csv(output_dir / "answer_metric_summary.csv", answer_summary)
    write_metric_csv(output_dir / "retrieval_metric_summary.csv", retrieval_summary)
    write_group_csv(output_dir / "source_summary.csv", source_summary)
    write_group_csv(output_dir / "topic_summary.csv", topic_summary)
    write_distribution_csv(output_dir / "answer_score_distributions.csv", answer_distributions)
    write_counter_csv(output_dir / "retrieval_grade_distribution.csv", retrieval_grade_distribution, "grade", "count")
    write_counter_csv(output_dir / "first_relevant_rank_distribution.csv", first_relevant_distribution, "first_relevant_rank", "count")

    write_latex_table(output_dir / "answer_metric_summary.tex", answer_summary, "Metric")
    write_latex_table(output_dir / "retrieval_metric_summary.tex", retrieval_summary, "Metric")

    make_horizontal_bar_pdf(
        figure_dir / "answer_metric_means.pdf",
        "Answer evaluation: mean score by metric",
        [(item["label"], item["mean"]) for item in answer_summary],
        max_value=5.0,
        x_label="Mean score on 0-5 scale",
    )
    make_horizontal_bar_pdf(
        figure_dir / "retrieval_metric_means.pdf",
        "Retrieval evaluation: mean value by metric",
        [(item["label"], item["mean"]) for item in retrieval_summary if item["mean"] is not None],
        max_value=1.0,
        x_label="Mean value on 0-1 scale",
    )
    make_horizontal_bar_pdf(
        figure_dir / "topic_overall_scores.pdf",
        "Mean overall answer score by topic",
        [(item["group"], item["answer_overall_mean"]) for item in topic_summary if item["n"] >= 2],
        max_value=5.0,
        x_label="Mean overall score on 0-5 scale",
    )

    print(f"Read {len(enriched_rows)} benchmark rows from {input_path}")
    print(f"Retrieval-evaluated rows: {len(retrieval_rows)}")
    print(f"Wrote analysis tables to {output_dir}")
    print(f"Wrote charts to {figure_dir}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyze the evaluated DORA RAG benchmark workbook.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="Evaluated workbook path.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for CSV, JSON, and TeX summaries.")
    parser.add_argument("--figure-dir", default=str(DEFAULT_FIGURE_DIR), help="Directory for PDF charts used by LaTeX.")
    return parser.parse_args()


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    row_iter = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(row_iter)]
    rows: list[dict[str, Any]] = []
    for row_values in row_iter:
        rows.append(dict(zip(headers, row_values)))
    return rows


def enrich_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for row in rows:
        copy = dict(row)
        copy["topic_normalized"] = normalize_topic(copy.get("QA Topic"))
        copy["source"] = clean_text(copy.get("source")).upper()
        copy["retrieval_judgments"] = parse_json_list(copy.get("Retrieval Chunk Judgments JSON"))
        enriched.append(copy)
    return enriched


def normalize_topic(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return "Unspecified"
    lower = re.sub(r"\s+", " ", text.lower())
    if "register of information" in lower:
        return "Register of information"
    if "third-party risk" in lower or "third party risk" in lower:
        return "ICT third-party risk management"
    if "incident" in lower:
        return "ICT-related incidents"
    if "risk management" in lower:
        return "ICT risk management"
    if "testing" in lower:
        return "Digital operational resilience testing"
    if "reporting template" in lower:
        return "Reporting templates"
    if "oversight" in lower:
        return "Oversight framework"
    if "scope" in lower:
        return "Scope"
    if "key function" in lower:
        return "Key functions"
    if "other dora" in lower:
        return "Other DORA topics"
    return text


def summarize_metrics(rows: list[dict[str, Any]], metrics: list[tuple[str, str]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for column, label in metrics:
        values = numeric_values(row.get(column) for row in rows)
        result.append(
            {
                "column": column,
                "label": label,
                "n": len(values),
                "mean": round(mean(values), 4) if values else None,
                "median": round(median(values), 4) if values else None,
                "min": round(min(values), 4) if values else None,
                "max": round(max(values), 4) if values else None,
                "std": round(pstdev(values), 4) if len(values) > 1 else 0.0 if values else None,
                "zero_count": sum(1 for value in values if value == 0),
                "high_count": sum(1 for value in values if value >= high_threshold(column)),
            }
        )
    return result


def high_threshold(column: str) -> float:
    if column.startswith("Retrieval"):
        return 0.8
    return 4.0


def summarize_by_group(rows: list[dict[str, Any]], group_column: str) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        group = clean_text(row.get(group_column)) or "Unspecified"
        groups[group].append(row)

    result: list[dict[str, Any]] = []
    for group, group_rows in groups.items():
        retrieval_rows = [row for row in group_rows if truthy(row.get("Retrieval Reference Evaluation Needed"))]
        result.append(
            {
                "group": group,
                "n": len(group_rows),
                "answer_overall_mean": rounded_mean(row.get("Answer Overall Score") for row in group_rows),
                "correctness_mean": rounded_mean(row.get("Answer Correctness Score") for row in group_rows),
                "faithfulness_mean": rounded_mean(row.get("Faithfulness Score") for row in group_rows),
                "completeness_mean": rounded_mean(row.get("Completeness Score") for row in group_rows),
                "citation_precision_mean": rounded_mean(row.get("Citation Precision Score") for row in group_rows),
                "citation_recall_mean": rounded_mean(row.get("Citation Recall Score") for row in group_rows),
                "retrieval_n": len(retrieval_rows),
                "precision_at_5_mean": rounded_mean(row.get("Retrieval Precision@5") for row in retrieval_rows),
                "mrr_mean": rounded_mean(row.get("Retrieval MRR") for row in retrieval_rows),
                "hit_at_5_mean": rounded_mean(row.get("Retrieval Hit@5") for row in retrieval_rows),
                "hit_at_10_mean": rounded_mean(row.get("Retrieval Hit@10") for row in retrieval_rows),
                "ndcg_at_5_mean": rounded_mean(row.get("Retrieval nDCG@5") for row in retrieval_rows),
            }
        )
    return sorted(result, key=lambda item: (-item["n"], item["group"]))


def score_distributions(rows: list[dict[str, Any]], metrics: list[tuple[str, str]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for column, label in metrics:
        counts = Counter(int(value) for value in numeric_values(row.get(column) for row in rows))
        item = {"metric": label}
        for score in range(0, 6):
            item[str(score)] = counts.get(score, 0)
        result.append(item)
    return result


def chunk_grade_distribution(rows: list[dict[str, Any]]) -> Counter[int]:
    counts: Counter[int] = Counter()
    for row in rows:
        for judgment in row.get("retrieval_judgments") or []:
            try:
                grade = int(judgment.get("relevance_grade"))
            except (TypeError, ValueError):
                continue
            counts[grade] += 1
    return counts


def first_relevant_rank(row: dict[str, Any]) -> str:
    ranks: list[int] = []
    for judgment in row.get("retrieval_judgments") or []:
        try:
            rank = int(judgment.get("rank"))
            grade = int(judgment.get("relevance_grade"))
        except (TypeError, ValueError):
            continue
        if grade >= 2:
            ranks.append(rank)
    if not ranks:
        return "none"
    return str(min(ranks))


def compute_correlations(rows: list[dict[str, Any]]) -> dict[str, Any]:
    pairs = [
        ("Retrieval Precision@5", "Answer Overall Score"),
        ("Retrieval MRR", "Answer Overall Score"),
        ("Retrieval Hit@5", "Answer Overall Score"),
        ("Retrieval nDCG@5", "Answer Overall Score"),
        ("Retrieval Hit@10", "Answer Overall Score"),
        ("Retrieval Precision@5", "Completeness Score"),
        ("Retrieval Hit@5", "Citation Recall Score"),
    ]
    result: dict[str, Any] = {}
    for left, right in pairs:
        values = []
        for row in rows:
            a = to_float(row.get(left))
            b = to_float(row.get(right))
            if a is not None and b is not None:
                values.append((a, b))
        result[f"{left} vs {right}"] = {
            "n": len(values),
            "pearson": round(pearson(values), 4) if len(values) >= 2 else None,
        }
    return result


def pearson(values: list[tuple[float, float]]) -> float:
    xs = [item[0] for item in values]
    ys = [item[1] for item in values]
    x_mean = mean(xs)
    y_mean = mean(ys)
    numerator = sum((x - x_mean) * (y - y_mean) for x, y in values)
    x_den = math.sqrt(sum((x - x_mean) ** 2 for x in xs))
    y_den = math.sqrt(sum((y - y_mean) ** 2 for y in ys))
    if x_den == 0 or y_den == 0:
        return 0.0
    return numerator / (x_den * y_den)


def numeric_values(values: Any) -> list[float]:
    result: list[float] = []
    for value in values:
        converted = to_float(value)
        if converted is not None:
            result.append(converted)
    return result


def rounded_mean(values: Any) -> float | None:
    nums = numeric_values(values)
    return round(mean(nums), 4) if nums else None


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = clean_text(value).lower()
    return text in {"true", "1", "yes"}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_json_list(value: Any) -> list[dict[str, Any]]:
    text = clean_text(value)
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list):
        return []
    return [item for item in parsed if isinstance(item, dict)]


def stringify_counter(counter: Counter[Any]) -> dict[str, int]:
    return {str(key): value for key, value in sorted(counter.items(), key=lambda item: str(item[0]))}


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_metric_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    write_csv(path, rows, ["label", "n", "mean", "median", "min", "max", "std", "zero_count", "high_count"])


def write_group_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = [
        "group",
        "n",
        "answer_overall_mean",
        "correctness_mean",
        "faithfulness_mean",
        "completeness_mean",
        "citation_precision_mean",
        "citation_recall_mean",
        "retrieval_n",
        "precision_at_5_mean",
        "mrr_mean",
        "hit_at_5_mean",
        "hit_at_10_mean",
        "ndcg_at_5_mean",
    ]
    write_csv(path, rows, fields)


def write_distribution_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    write_csv(path, rows, ["metric", "0", "1", "2", "3", "4", "5"])


def write_counter_csv(path: Path, counter: Counter[Any], key_name: str, value_name: str) -> None:
    rows = [{key_name: key, value_name: value} for key, value in sorted(counter.items(), key=lambda item: str(item[0]))]
    write_csv(path, rows, [key_name, value_name])


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field) for field in fields})


def write_latex_table(path: Path, rows: list[dict[str, Any]], first_column_name: str) -> None:
    lines = [
        "\\begin{tabular}{lrrrrrrr}",
        "\\hline",
        f"{first_column_name} & N & Mean & Median & Min & Max & Std. & 0 count \\\\",
        "\\hline",
    ]
    for row in rows:
        lines.append(
            " & ".join(
                [
                    latex_escape(str(row["label"])),
                    str(row["n"]),
                    format_number(row["mean"]),
                    format_number(row["median"]),
                    format_number(row["min"]),
                    format_number(row["max"]),
                    format_number(row["std"]),
                    str(row["zero_count"]),
                ]
            )
            + " \\\\"
        )
    lines.extend(["\\hline", "\\end{tabular}", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def latex_escape(text: str) -> str:
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(replacements.get(char, char) for char in text)


def format_number(value: Any) -> str:
    if value is None:
        return "-"
    number = float(value)
    if abs(number - round(number)) < 0.000001:
        return f"{number:.0f}"
    return f"{number:.2f}"


def make_horizontal_bar_pdf(path: Path, title: str, data: list[tuple[str, float | None]], max_value: float, x_label: str) -> None:
    clean_data = [(label, value) for label, value in data if value is not None]
    clean_data = sorted(clean_data, key=lambda item: item[1])
    width = 760
    row_height = 31
    top_margin = 58
    bottom_margin = 50
    left_margin = 255
    right_margin = 55
    height = max(270, top_margin + bottom_margin + row_height * len(clean_data))
    chart_width = width - left_margin - right_margin
    content: list[str] = []

    content.append("1 1 1 rg 0 0 {0} {1} re f".format(width, height))
    text(content, 28, height - 30, title, 15)
    text(content, left_margin, 28, x_label, 9)

    for i in range(6):
        tick_value = max_value * i / 5
        x = left_margin + chart_width * tick_value / max_value
        content.append("0.86 0.86 0.86 RG 0.5 w {0:.2f} {1:.2f} m {0:.2f} {2:.2f} l S".format(x, bottom_margin, height - top_margin + 10))
        text(content, x - 8, bottom_margin - 16, f"{tick_value:.1f}", 7)

    palette = [(0.10, 0.33, 0.57), (0.05, 0.50, 0.45), (0.69, 0.32, 0.16), (0.33, 0.40, 0.75)]
    for idx, (label, value) in enumerate(clean_data):
        y = bottom_margin + idx * row_height + 7
        bar_width = 0 if max_value == 0 else chart_width * max(0, min(value, max_value)) / max_value
        r, g, b = palette[idx % len(palette)]
        text(content, 28, y + 4, shorten(label, 42), 8)
        content.append(f"{r:.3f} {g:.3f} {b:.3f} rg {left_margin:.2f} {y:.2f} {bar_width:.2f} 16 re f")
        text(content, left_margin + bar_width + 6, y + 4, f"{value:.2f}", 8)

    content.append("0 0 0 RG 0.8 w {0} {1} m {0} {2} l S".format(left_margin, bottom_margin, height - top_margin + 16))
    write_pdf(path, width, height, "\n".join(content))


def shorten(text_value: str, limit: int) -> str:
    if len(text_value) <= limit:
        return text_value
    return text_value[: limit - 3].rstrip() + "..."


def text(content: list[str], x: float, y: float, value: str, size: int) -> None:
    safe = pdf_escape(value)
    content.append(f"0 0 0 rg BT /F1 {size} Tf {x:.2f} {y:.2f} Td ({safe}) Tj ET")


def pdf_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def write_pdf(path: Path, width: int, height: int, content: str) -> None:
    stream = content.encode("latin-1", errors="replace")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>".encode("ascii"),
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    chunks = [b"%PDF-1.4\n"]
    offsets: list[int] = []
    for index, obj in enumerate(objects, start=1):
        offsets.append(sum(len(chunk) for chunk in chunks))
        chunks.append(f"{index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n")
    xref_offset = sum(len(chunk) for chunk in chunks)
    chunks.append(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii"))
    for offset in offsets:
        chunks.append(f"{offset:010d} 00000 n \n".encode("ascii"))
    chunks.append(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    path.write_bytes(b"".join(chunks))


if __name__ == "__main__":
    main()
